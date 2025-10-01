import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import os
from collections import defaultdict

# Cores
TIPO_CORES = ["FFC7CE", "C6EFCE", "FFEB9C", "BDD7EE", "F4CCCC"]
LINHA_ALTERNADA = ["FFFFFF", "F2F2F2"]

def salvar_producao_diaria(dicionario_produtos, data=datetime.today().date(), nome_arquivo="Loja012_2025.xlsx"):
    data_str = data.strftime("%d-%m-%Y")

    # Abrir ou criar arquivo
    if os.path.exists(nome_arquivo):
        wb = openpyxl.load_workbook(nome_arquivo)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # Abrir ou criar aba
    if data_str in wb.sheetnames:
        ws = wb[data_str]
    else:
        ws = wb.create_sheet(title=data_str)

    # Preparar dados
    existentes = {}
    for produto in dicionario_produtos.values():
        nome = produto.getnome()
        tipo = produto.gettipo()
        quantidade = produto.getquantidade()
        if nome in existentes:
            existentes[nome]['quantidade'] += quantidade
        else:
            existentes[nome] = {'tipo': tipo, 'quantidade': quantidade}

    # Agrupar por tipo e ordenar decrescente
    por_tipo = defaultdict(list)
    for nome, info in existentes.items():
        por_tipo[info['tipo']].append((nome, info['quantidade']))
    for tipo in por_tipo:
        por_tipo[tipo].sort(key=lambda x: x[1], reverse=True)

    # Configurações da tabela
    headers = ["Nome", "Tipo", "Quantidade"]
    start_row = 3
    start_col = 5  # Coluna E

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabeçalho
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + col_offset, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col + col_offset)].width = [20, 15, 12][col_offset]

    # Inserir dados
    current_row = start_row + 1
    total_geral = 0
    cor_index = 0
    for tipo in sorted(por_tipo.keys()):
        cor_fill = PatternFill(start_color=TIPO_CORES[cor_index % len(TIPO_CORES)],
                               end_color=TIPO_CORES[cor_index % len(TIPO_CORES)],
                               fill_type="solid")
        total_tipo = 0
        for i, (nome, quantidade) in enumerate(por_tipo[tipo]):
            for col_offset, value in enumerate([nome, tipo, quantidade]):
                cell = ws.cell(row=current_row, column=start_col + col_offset, value=value)
                fill = PatternFill(start_color=LINHA_ALTERNADA[i % 2], end_color=LINHA_ALTERNADA[i % 2], fill_type="solid")
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 18  # Altura da linha
            total_tipo += quantidade
            total_geral += quantidade
            current_row += 1

        # Total por tipo
        for col_offset, value in enumerate([f"Total {tipo}", "", total_tipo]):
            cell = ws.cell(row=current_row, column=start_col + col_offset, value=value)
            cell.font = Font(bold=True, size=11)
            cell.fill = cor_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_offset != 2 else "right", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        cor_index += 1

    # Total geral
    for col_offset, value in enumerate(["TOTAL GERAL", "", total_geral]):
        cell = ws.cell(row=current_row, column=start_col + col_offset, value=value)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col_offset != 2 else "right", vertical="center")
    ws.row_dimensions[current_row].height = 22

    wb.save(nome_arquivo)
    print(f"✅ Planilha atualizada na aba '{data_str}' a partir de E3 → {nome_arquivo}")

